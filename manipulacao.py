"""
Projeto 2
Descrição da atividade: Escreva um programa em Python que crie planilhas dentro do arquivo orcamento.xls que se encontra na pasta planilhas que vocˆe criou. As planilhas
devem ter os nomes a seguir
• receitas
• despesas
• resultado
Autor: Daniel Francisco Bristot
05/09/2022

"""

print ("\nIniciando o programa")

#importando o módulo openpyxl para trabalhar com planilhas
from openpyxl import load_workbook

#abrir a planilha
wb=load_workbook("orcamento.xlsx")

#renomeando a aba sheet 1
ws1 = wb.active
ws1.title = 'receitas'

#criando as outras 2 abas
ws2=wb.create_sheet("despesas")
ws3=wb.create_sheet("resultado")

#salvando o arquivo
wb.save("orcamento.xlsx")

print ("\nFinalizado a manipulação da planilha.")
